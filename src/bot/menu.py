from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes

from src.database.database import SessionLocal
from src.database.crud import (
    obter_resumo_mes,  obter_analise_categorias, 
    listar_metas, verificar_meta_categoria, Transacao, Renda
    )
from datetime import date, datetime
import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def gerar_botoes_meses():
    hoje = date.today()
    meses_nomes = {
        1:"Janeiro", 2:"Fevereiro", 3:"Março", 4:"Abril", 
        5:"Maio", 6:"Junho", 7:"Julho", 8:"Agosto", 
        9:"Setembro", 10:"Outubro", 11:"Novembro", 12:"Dezembro"
    }
    
    botoes = []
    mes_atual = hoje.month
    ano_atual = hoje.year
    
    for _ in range(4): 
        nome_mes = meses_nomes[mes_atual]
        texto_botao = f"📅 {nome_mes}/{ano_atual}"
        callback = f"btn_rel_mes_{mes_atual:02d}_{ano_atual}" 
        botoes.append([InlineKeyboardButton(texto_botao, callback_data=callback)])
        
        mes_atual -= 1
        if mes_atual == 0:
            mes_atual = 12
            ano_atual -= 1
            
    botoes.append([InlineKeyboardButton("📚 Tudo", callback_data="btn_rel_tudo")])
    botoes.append([InlineKeyboardButton("🔙 Voltar", callback_data="btn_voltar")])
    
    return botoes

def gerar_botoes_tipo_renda():
    return [
        [
            InlineKeyboardButton("💰 Salário", callback_data="renda_tipo_Salário"),
            InlineKeyboardButton("🏦 Adiantamento", callback_data="renda_tipo_Adiantamento")
        ],
        [
            InlineKeyboardButton("🎁 Benefício", callback_data="renda_tipo_Benefício"),
            InlineKeyboardButton("➕ Extra", callback_data="renda_tipo_Extra")
        ],
        [InlineKeyboardButton("🔙 Voltar ao Menu", callback_data="renda_cancelar")]
    ]

async def comando_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    
    teclado = [
            [InlineKeyboardButton("💸 Nova Renda", callback_data="renda_start"),
            InlineKeyboardButton("📊 Saldo", callback_data="btn_saldo")], 

            [InlineKeyboardButton("📋 Extrato", callback_data="btn_extrato"),
            InlineKeyboardButton("📈 Categorias", callback_data="btn_analise")],
            
            [InlineKeyboardButton("🎯 Minhas Metas", callback_data="btn_metas")],
            [InlineKeyboardButton("📁 Relatório Excel", callback_data="btn_relatorio")]
        ]
    
    reply_markup = InlineKeyboardMarkup(teclado)
    
    if update.callback_query:
        await update.callback_query.message.edit_text("Painel de Controle FinAI 🤖\nEscolha uma opção:", reply_markup=reply_markup)
    else:
        await update.message.reply_text("Painel de Controle FinAI 🤖\nEscolha uma opção:", reply_markup=reply_markup)


async def processar_cliques_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer() 
    
    escolha = query.data
    chat_id = update.effective_chat.id
    
    if escolha == "btn_saldo":
        db = SessionLocal()
        try:
            resumo = obter_resumo_mes(db, chat_id)
            
            receitas_totais = resumo["receitas_totais"]
            despesas_mes = resumo["despesas_mes"]
            saldo_bancario = resumo["saldo_bancario"]
            
            rec_fmt = f"{receitas_totais:.2f}".replace('.', ',')
            desp_mes_fmt = f"{despesas_mes:.2f}".replace('.', ',')
            saldo_fmt = f"{saldo_bancario:.2f}".replace('.', ',')
            
            texto_saldo = (
                "🏦 *Conta Bancária (Saldo Real)*\n"
                f"💰 *Disponível:* R$ {saldo_fmt}\n\n"
                "━━━━━━━━━━━━━━━━\n"
                "📊 *Controle deste Mês*\n"
                f"📉 Gastos do Mês: R$ {desp_mes_fmt}\n"
                "━━━━━━━━━━━━━━━━\n"
            )

            if saldo_bancario > 0:
                texto_saldo += "✅ O desafio de economia segue firme!"
            elif saldo_bancario == 0:
                texto_saldo += "⚖️ Conta zerada. Atenção aos próximos gastos."
            else:
                texto_saldo += "🚨 Saldo negativo! Você está no cheque especial."
            
            teclado_voltar = [[InlineKeyboardButton("🔙 Voltar ao Menu", callback_data="btn_voltar")]]
            reply_markup = InlineKeyboardMarkup(teclado_voltar)
            
            await query.edit_message_text(texto_saldo, parse_mode="Markdown", reply_markup=reply_markup)
            
        except Exception as e:
            await query.edit_message_text(f"❌ Erro ao buscar saldo: {e}")
        finally:
            db.close()

    elif escolha == "btn_voltar":
        await comando_menu(update, context)
        
    elif escolha == "btn_analise":
        db = SessionLocal()
        try:
            gastos = obter_analise_categorias(db, chat_id)
            
            if not gastos:
                texto_analise = "📊 Ainda não existem despesas registradas para análise."
            else:
                total_gastos = sum(item["total"] for item in gastos)
                
                texto_analise = "📊 *Análise de Gastos por Categoria*\n━━━━━━━━━━━━━━━━\n"
                
                for item in gastos:
                    categoria = item["categoria"]
                    valor = item["total"]
                    
                    percentual = (valor / total_gastos) * 100
                    valor_formatado = f"{valor:.2f}".replace('.', ',')
                    
                    tamanho_barra = int(percentual / 10)
                    barra = "🟩" * tamanho_barra + "⬜" * (10 - tamanho_barra)
                    
                    texto_analise += f"*{categoria}* - {percentual:.1f}%\n"
                    texto_analise += f"{barra} R$ {valor_formatado}\n\n"
                
                texto_analise += "━━━━━━━━━━━━━━━━\n"
                texto_analise += f"🔴 *Total Gasto:* R$ {f'{total_gastos:.2f}'.replace('.', ',')}"
            
            teclado_voltar = [[InlineKeyboardButton("🔙 Voltar ao Menu", callback_data="btn_voltar")]]
            reply_markup = InlineKeyboardMarkup(teclado_voltar)
            
            await query.edit_message_text(texto_analise, parse_mode="Markdown", reply_markup=reply_markup)
            
        except Exception as e:
            await query.edit_message_text(f"❌ Erro ao gerar análise: {e}")
        finally:
            db.close()

    elif escolha == "btn_extrato":
        db = SessionLocal()
        try:
            saidas = db.query(Transacao).filter(Transacao.chat_id == chat_id).order_by(Transacao.id.desc()).limit(10).all()
            entradas = db.query(Renda).filter(Renda.chat_id == chat_id).order_by(Renda.id.desc()).limit(5).all()
            
            movimentacoes = []
            hoje = date.today()
            
            for s in saidas:
                data_obj = s.data if s.data else datetime.combine(hoje, datetime.min.time())
                movimentacoes.append({
                    "tipo": "saida", "valor": s.valor, "categoria": s.categoria, 
                    "descricao": s.descricao, "data": data_obj, "id_num": s.id
                })
                
            for e in entradas:
                try:
                    data_obj = datetime(hoje.year, hoje.month, int(e.dia_recebimento))
                except:
                    data_obj = datetime.combine(hoje, datetime.min.time())
                    
                movimentacoes.append({
                    "tipo": "entrada", "valor": e.valor, "categoria": "Receita/Renda", 
                    "descricao": e.descricao, "data": data_obj, "id_num": e.id
                })
                
            movimentacoes.sort(key=lambda x: (x["data"], 1 if x["tipo"] == "entrada" else 0, x["id_num"]), reverse=True)
            
            ultimas = movimentacoes[:5]
            
            if not ultimas:
                texto_ultimos = "🔍 Nenhuma movimentação recente encontrada."
            else:
                texto_ultimos = "🔍 *Últimas 5 Movimentações*\n━━━━━━━━━━━━━━━━\n"
                for m in ultimas:
                    icone = "🔴" if m["tipo"] == "saida" else "🟢"
                    data_formatada = m["data"].strftime("%d/%m")
                    valor_formatado = f"{m['valor']:.2f}".replace('.', ',')
                    
                    texto_ultimos += f"{icone} *{data_formatada}* | {m['categoria']}\n"
                    texto_ultimos += f"    └ R$ {valor_formatado} ({m['descricao']})\n\n"
                    
            teclado_voltar = [[InlineKeyboardButton("🔙 Voltar ao Menu", callback_data="btn_voltar")]]
            await query.edit_message_text(texto_ultimos, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(teclado_voltar))
            
        except Exception as e:
            await query.edit_message_text(f"❌ Erro ao buscar extrato: {e}")
        finally:
            db.close()

    elif escolha == "btn_relatorio":
        await query.edit_message_text(
            "📅 *Selecione o mês do relatório:*", 
            reply_markup=InlineKeyboardMarkup(gerar_botoes_meses()), 
            parse_mode="Markdown"
        )

    elif escolha.startswith("btn_rel_"):
        await query.edit_message_text("⏳ *Filtrando dados e gerando Excel...*", parse_mode="Markdown")
        
        mes_filtro = None
        ano_filtro = None
        nome_arquivo_periodo = "Completo"
        
        if escolha.startswith("btn_rel_mes_"):
            partes = escolha.split("_") 
            mes_filtro = int(partes[3])
            ano_filtro = int(partes[4])
            nome_arquivo_periodo = f"{mes_filtro:02d}_{ano_filtro}"

        db = SessionLocal()
        try:            
            transacoes_banco = db.query(Transacao).filter(Transacao.chat_id == chat_id).all()
            rendas_fixas = db.query(Renda).filter(Renda.chat_id == chat_id).all()
            
            dados_unificados = []
            hoje = date.today()
            
            for t in transacoes_banco:
                data_item = t.data.date() if getattr(t, 'data', None) else hoje
                
                if t.tipo == "saida":
                    dados_unificados.append({
                        "Data": data_item, 
                        "Tipo": "Saída", 
                        "Categoria": t.categoria,
                        "Descrição": t.descricao, 
                        "Valor": t.valor * -1 
                    })
                else: 
                    dados_unificados.append({
                        "Data": data_item, 
                        "Tipo": "Entrada", 
                        "Categoria": t.categoria,
                        "Descrição": t.descricao, 
                        "Valor": t.valor 
                    })
                
            for e in rendas_fixas:
                try:
                    ano_renda = ano_filtro if ano_filtro else hoje.year
                    mes_renda = mes_filtro if mes_filtro else hoje.month
                    
                    if e.dia_recebimento:
                        data_item = date(ano_renda, mes_renda, int(e.dia_recebimento))
                    else:
                        data_item = date(ano_renda, mes_renda, hoje.day)
                except (ValueError, TypeError):
                    data_item = hoje
                    
                dados_unificados.append({
                    "Data": data_item, 
                    "Tipo": "Entrada Fixa", 
                    "Categoria": "Salário/Benefício",
                    "Descrição": e.descricao, 
                    "Valor": e.valor
                })
            if mes_filtro and ano_filtro:
                dados_unificados = [
                    item for item in dados_unificados 
                    if item["Data"].month == mes_filtro and item["Data"].year == ano_filtro
                ]

            if not dados_unificados:
                teclado_voltar = [[InlineKeyboardButton("🔙 Voltar", callback_data="btn_relatorio")]]
                await query.edit_message_text(
                    "📭 Nenhuma movimentação encontrada neste mês.", 
                    reply_markup=InlineKeyboardMarkup(teclado_voltar)
                )
                return

            df = pd.DataFrame(dados_unificados)
            df = df.sort_values(by="Data", ascending=False)
            
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Relatório FinAI')
                workbook = writer.book
                worksheet = writer.sheets['Relatório FinAI']

                
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                verde_suave = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                vermelho_suave = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                cinza_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                center_alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin'))

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border

                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    data_cell = row[0]
                    tipo_cell = row[1]
                    valor_cell = row[4]

                    data_cell.number_format = 'DD/MM/YYYY'
                    data_cell.alignment = center_alignment
                    tipo_cell.alignment = center_alignment

                    if "Saída" in str(tipo_cell.value):
                        tipo_cell.fill = vermelho_suave
                        valor_cell.font = Font(color="9C0006", bold=True) 
                    else:
                        tipo_cell.fill = verde_suave
                        valor_cell.font = Font(color="006100", bold=True) 
                    
                    valor_cell.number_format = '"R$" #,##0.00'
                    
                    for cell in row:
                        cell.border = thin_border

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                length = len(str(cell.value))
                                if length > max_length:
                                    max_length = length
                        except:
                            pass
                    
                    if column_letter == 'E':
                        worksheet.column_dimensions[column_letter].width = max(max_length + 6, 18)
                    else:
                        worksheet.column_dimensions[column_letter].width = max_length + 4

                total_row = worksheet.max_row + 1
                saldo = df['Valor'].sum()
                
                for col_idx in range(1, 6):
                    cell = worksheet.cell(row=total_row, column=col_idx)
                    cell.fill = cinza_total
                    cell.border = thin_border
                
                label_cell = worksheet.cell(row=total_row, column=4, value="SALDO TOTAL:")
                label_cell.font = Font(bold=True)
                label_cell.alignment = Alignment(horizontal="right")
                
                saldo_cell = worksheet.cell(row=total_row, column=5, value=saldo)
                saldo_cell.font = Font(bold=True, color="000000")
                saldo_cell.number_format = '"R$" #,##0.00'

            buffer.seek(0)
            
            nome_periodo = escolha.replace("btn_rel_", "")
            arquivo_nome = f"Relatorio_FinAI_{nome_periodo}dias.xlsx" if nome_periodo != "tudo" else "Relatorio_FinAI_Completo.xlsx"
            
            await context.bot.send_document(
                chat_id=chat_id,
                document=buffer,
                filename=f"Relatorio_FinAI_{nome_arquivo_periodo}.xlsx",
                caption="✅ *Seu relatório está pronto!*",
                parse_mode="Markdown"
            )
            
            teclado_voltar = [[InlineKeyboardButton("🔙 Voltar ao Menu", callback_data="btn_voltar")]]
            await query.edit_message_text("📈 Relatório enviado com sucesso!", reply_markup=InlineKeyboardMarkup(teclado_voltar))
            
        except Exception as e:
            await query.edit_message_text(f"❌ Erro ao gerar Excel: {e}")
        finally:
            db.close()

    elif escolha == "btn_metas":
        await query.edit_message_text("⏳ *Buscando suas metas...*", parse_mode="Markdown")
        
        db = SessionLocal()
        try:
            metas_cadastradas = listar_metas(db, chat_id)
            
            if not metas_cadastradas:
                teclado_voltar = [[InlineKeyboardButton("🔙 Voltar ao Menu", callback_data="btn_voltar")]]
                await query.edit_message_text(
                    "📭 *Nenhuma meta definida ainda.*\nUse o comando `/meta Categoria Valor` para criar.", 
                    reply_markup=InlineKeyboardMarkup(teclado_voltar),
                    parse_mode="Markdown"
                )
                return

            mensagem = "🎯 *Suas Metas do Mês*\n━━━━━━━━━━━━━━━━\n"
            
            for meta in metas_cadastradas:
                info = verificar_meta_categoria(db, chat_id, meta.categoria)
                if info:
                    percentual = info['percentual']
                    gasto_fmt = f"{info['gasto']:.2f}".replace('.', ',')
                    limite_fmt = f"{info['limite']:.2f}".replace('.', ',')
                    restante_fmt = f"{abs(info['restante']):.2f}".replace('.', ',')
                    
                    tamanho_barra = min(int(percentual / 10), 10)
                    
                    if percentual > 100:
                        barra = "🟥" * 10
                        icone = "🚨"
                        status_sobra = f"🛑 *Estourou R$ {restante_fmt}*"
                    elif percentual >= 80:
                        barra = "🟧" * tamanho_barra + "⬜" * (10 - tamanho_barra)
                        icone = "⚠️"
                        status_sobra = f"🟡 *Cuidado! Sobra R$ {restante_fmt}*"
                    else:
                        barra = "🟩" * tamanho_barra + "⬜" * (10 - tamanho_barra)
                        icone = "✅"
                        status_sobra = f"🟢 *Livre: R$ {restante_fmt}*"
                    
                    mensagem += f"{icone} *{meta.categoria}* ({percentual:.1f}%)\n"
                    mensagem += f"{barra}\n"
                    mensagem += f"💰 R$ {gasto_fmt} / R$ {limite_fmt}\n"
                    mensagem += f"↳ {status_sobra}\n\n"

            teclado_voltar = [[InlineKeyboardButton("🔙 Voltar ao Menu", callback_data="btn_voltar")]]
            await query.edit_message_text(mensagem, reply_markup=InlineKeyboardMarkup(teclado_voltar), parse_mode="Markdown")

        except Exception as e:
            await query.edit_message_text(f"❌ Erro ao buscar metas: {e}")
        finally:
            db.close()